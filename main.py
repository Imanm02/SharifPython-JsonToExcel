import pandas as pd
import json
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

with open("9090.json", 'r', encoding='utf-8') as f:
    data = json.load(f)

messenger_type_mapping = {
    'bale': 'بله',
    'eitaa': 'ایتا',
    'rubika': 'روبیکا',
    'gap': 'گپ',
    'soroush': 'سروش',
    'igap': 'آی‌گپ'
}

students_data = []

for course in data['courses']:
    for group in course['current_group']:
        for student in group['students']:
            student_data = {
                'نام': student['name'],
                'نام خانوادگی': student['surname'],
                'جنسیت': 'آقا' if student['gender'] == 'male' else 'خانم',
                'ایمیل': student['email'],
                'شماره تلفن': student['mobile_number'],
                'کد ملی': student['national_code'],
                'شماره تلفن ثابت': student['phone_number'],
                'تاریخ آپدیت اطلاعات': student['updated_at'],
                'وضعیت': 'ثبت‌نام' if student['pivot']['status'] == 'pending' else 'کردیت',
            }

            if student['extra'] and isinstance(student['extra'], dict):
                for key, value in student['extra'].items():
                    if key == 'telegram_number':
                        student_data['شماره تلگرام'] = value
                    elif key == 'whatsapp_number':
                        student_data['شماره واتس‌آپ'] = value
                    elif key == 'internal_messenger_type':
                        student_data['پیام‌رسان داخلی مورد استفاده'] = messenger_type_mapping.get(value, value)
                    elif key == 'internal_messenger_number':
                        student_data['شماره تلفن پیام‌رسان داخلی'] = value
                    elif key == 'in_person_classes':
                        student_data['وضعیت حضور در کلاس'] = 'حضوری' if value == True else 'مجازی'

            students_data.append(student_data)

df = pd.DataFrame(students_data)

df.to_excel('9090.xlsx', index=False)

wb = openpyxl.load_workbook('9090.xlsx')
sheet = wb.active

font = Font(name='Vazirmatn')
alignment = Alignment(horizontal='center', vertical='center')
light_yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
light_green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

for row in sheet:
    for cell in row:
        cell.font = font
        cell.alignment = alignment
        cell.number_format = '@'
        if cell.row == 1:
            cell.fill = light_yellow_fill
        else:
            cell.fill = light_green_fill

for column_cells in sheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

wb.save('9090.xlsx')
